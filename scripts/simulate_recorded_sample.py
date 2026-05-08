from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from huolala_quote_tool.excel_model import (
    create_output_path,
    detect_workbook,
    iter_data_rows,
    load_vehicle_rules,
    mark_failure,
    numeric_or_original,
    writable_cell,
    write_success,
)
from huolala_quote_tool.parsers import parse_fixed_price, parse_total_distance


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Use a recorded workbook as simulated Huolala page output.")
    parser.add_argument("source", type=Path, help="Workbook containing recorded distance and quote values.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path.cwd() / "simulation_output",
        help="Directory for the simulated result workbook.",
    )
    parser.add_argument("--limit", type=int, default=0, help="Maximum data rows to process; 0 means all rows.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rules = load_vehicle_rules()
    detection = detect_workbook(args.source, rules=rules)
    if not detection.can_start:
        raise SystemExit("字段识别不完整，无法模拟。")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    output_path = args.output_dir / create_output_path(args.source).name

    recorded_wb = load_workbook(args.source, data_only=False)
    recorded_ws = recorded_wb[detection.sheet_name]
    output_wb = load_workbook(args.source, data_only=False)
    output_ws = output_wb[detection.sheet_name]

    rows = iter_data_rows(output_ws, detection)
    if args.limit > 0:
        rows = rows[: args.limit]

    distance_col = detection.required_columns["distance"]
    assert distance_col is not None

    # Start from a blank quote state to mimic an unquoted input workbook.
    for row_idx in rows:
        writable_cell(output_ws, row_idx, distance_col.index).value = None
        writable_cell(output_ws, row_idx, distance_col.index).comment = None
        for vehicle_col in detection.vehicle_columns:
            cell = writable_cell(output_ws, row_idx, vehicle_col.index)
            cell.value = None
            cell.comment = None

    processed = 0
    failed = 0
    mismatches: list[str] = []

    for row_idx in rows:
        recorded_distance = recorded_ws.cell(row=row_idx, column=distance_col.index).value
        distance_cell = writable_cell(output_ws, row_idx, distance_col.index)
        if recorded_distance in (None, ""):
            mark_failure(distance_cell, "录制样本没有总里程")
            failed += 1
        else:
            distance_text = f"总里程{recorded_distance}公里"
            parsed_distance = parse_total_distance(distance_text)
            write_success(distance_cell, numeric_or_original(parsed_distance))

        for vehicle_col in detection.vehicle_columns:
            recorded_price = recorded_ws.cell(row=row_idx, column=vehicle_col.index).value
            price_cell = writable_cell(output_ws, row_idx, vehicle_col.index)
            if recorded_price in (None, ""):
                mark_failure(price_cell, "录制样本没有运费一口价")
                failed += 1
                continue
            page_text = f"运费一口价 {recorded_price}元\n总计 0.01元"
            parsed_price = parse_fixed_price(page_text)
            write_success(price_cell, numeric_or_original(parsed_price))

            if float(price_cell.value) != float(recorded_price):
                mismatches.append(f"第 {row_idx} 行 {vehicle_col.header}: {price_cell.value} != {recorded_price}")

        output_wb.save(output_path)
        processed += 1

    output_wb.save(output_path)

    print(f"source={args.source}")
    print(f"output={output_path}")
    print(f"sheet={detection.sheet_name}")
    print(f"header_row={detection.header_row}")
    print(f"rows_processed={processed}")
    print(f"vehicle_columns={len(detection.vehicle_columns)}")
    for vehicle_col in detection.vehicle_columns:
        assert vehicle_col.rule is not None
        print(f"vehicle={vehicle_col.header} -> {vehicle_col.rule.car_length} / {'、'.join(vehicle_col.rule.vehicle_requirements)}")
    print(f"failures_marked={failed}")
    print(f"mismatches={len(mismatches)}")
    for item in mismatches[:10]:
        print(f"mismatch={item}")


if __name__ == "__main__":
    main()
