from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .paths import config_dir


FAILURE_PREFIX = "货拉拉报价失败: "

ROLE_ALIASES = {
    "supplier": ["供应商名称", "供应商", "承运商", "客户名称", "客户"],
    "origin": ["发货地址", "发货地", "发货详细地址", "提货地址", "起运地址", "装货地址"],
    "destination": ["到货地址", "收货地址", "到货地", "目的地", "卸货地址", "送货地址"],
    "distance": ["距离", "总里程", "里程", "公里数", "运输距离"],
}

ROLE_LABELS = {
    "supplier": "供应商名称",
    "origin": "发货地址",
    "destination": "到货地址",
    "distance": "距离",
}

VEHICLE_HINT_RE = re.compile(
    r"(车|车型|车长|厢|箱|飞翼|平板|高栏|冷藏|栏板|货车|"
    r"(?:4|6|7|9|13|17)\s*(?:\.|米)?\s*(?:2|6|8|5)?)"
)


def default_rules_path() -> Path:
    rule_dir = config_dir()
    csv_path = rule_dir / "vehicle_rules.csv"
    return csv_path if csv_path.exists() else rule_dir / "vehicle_rules.json"


def _split_rule_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"[,，;；、|]\s*", text) if item.strip()]


def _enabled_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = normalize_text(value)
    return text not in {"0", "false", "no", "n", "否", "停用", "禁用", "disabled"}


def _load_rule_rows_from_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(row) for row in csv.DictReader(f) if any(row.values())]


def _load_rule_rows_from_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [display_text(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if any(not is_blank(value) for value in item.values()):
            rows.append(item)
    return rows


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    replacements = {
        "＆": "&",
        "　": " ",
        "（": "(",
        "）": ")",
        "：": ":",
        "；": ";",
        "，": ",",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return re.sub(r"\s+", "", text).strip().lower()


def display_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


@dataclass(frozen=True)
class VehicleRule:
    name: str
    excel_headers: tuple[str, ...]
    car_length: str
    vehicle_requirements: tuple[str, ...]
    header_regex: tuple[str, ...] = ()
    enabled: bool = True

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "VehicleRule":
        headers = _split_rule_values(data.get("excel_headers", data.get("Excel表头关键词", [])))
        requirements = _split_rule_values(data.get("vehicle_requirements", data.get("网页车型要求", [])))
        regex_values = _split_rule_values(data.get("header_regex", data.get("表头正则", [])))
        car_length = str(data.get("car_length", data.get("网页车长", ""))).strip()
        name = str(data.get("name", data.get("规则名称", ""))).strip()
        if not name:
            name = " / ".join(item for item in (headers[0] if headers else "", car_length) if item)
        return cls(
            name=name,
            excel_headers=tuple(headers),
            car_length=car_length,
            vehicle_requirements=tuple(requirements),
            header_regex=tuple(regex_values),
            enabled=_enabled_value(data.get("enabled", data.get("启用", True))),
        )

    def match_score(self, header: Any) -> float:
        normalized = normalize_text(header)
        if not normalized:
            return 0.0

        for alias in self.excel_headers:
            alias_normalized = normalize_text(alias)
            if normalized == alias_normalized:
                return 1.0
            if alias_normalized and alias_normalized in normalized:
                return 0.95

        raw = display_text(header)
        for pattern in self.header_regex:
            if re.search(pattern, raw, flags=re.IGNORECASE):
                return 0.9
        return 0.0


@dataclass
class DetectedColumn:
    role: str
    header: str
    index: int
    confidence: float
    rule: VehicleRule | None = None
    message: str = ""

    @property
    def letter(self) -> str:
        return get_column_letter(self.index)

    @property
    def label(self) -> str:
        if self.role == "vehicle":
            return self.rule.name if self.rule else "未匹配车型规则"
        return ROLE_LABELS.get(self.role, self.role)


@dataclass
class DetectionResult:
    workbook_path: Path
    sheet_name: str
    header_row: int
    required_columns: dict[str, DetectedColumn | None]
    vehicle_columns: list[DetectedColumn] = field(default_factory=list)
    unmatched_vehicle_headers: list[DetectedColumn] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    max_row: int = 0

    @property
    def missing_required_roles(self) -> list[str]:
        return [ROLE_LABELS[role] for role, col in self.required_columns.items() if col is None]

    @property
    def can_start(self) -> bool:
        return not self.missing_required_roles and bool(self.vehicle_columns)


def load_vehicle_rules(path: str | Path | None = None) -> list[VehicleRule]:
    rules_path = Path(path) if path else default_rules_path()
    if rules_path.suffix.lower() in {".csv", ".txt"}:
        data = _load_rule_rows_from_csv(rules_path)
    elif rules_path.suffix.lower() in {".xlsx", ".xlsm"}:
        data = _load_rule_rows_from_workbook(rules_path)
    else:
        with rules_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    return [rule for rule in (VehicleRule.from_dict(item) for item in data) if rule.enabled]


def role_match_score(header: Any, role: str) -> float:
    normalized = normalize_text(header)
    if not normalized:
        return 0.0
    for alias in ROLE_ALIASES[role]:
        alias_normalized = normalize_text(alias)
        if normalized == alias_normalized:
            return 1.0
        if alias_normalized and alias_normalized in normalized:
            return 0.88
    return 0.0


def best_vehicle_rule(header: Any, rules: Iterable[VehicleRule]) -> tuple[VehicleRule | None, float]:
    best_rule = None
    best_score = 0.0
    for rule in rules:
        score = rule.match_score(header)
        if score > best_score:
            best_rule = rule
            best_score = score
    return best_rule, best_score


def header_candidates(ws: Worksheet, row_idx: int, col_idx: int) -> list[str]:
    header = display_text(ws.cell(row_idx, col_idx).value)
    if not header:
        return []

    candidates = [header]
    if row_idx > 1:
        parent = display_text(ws.cell(row_idx - 1, col_idx).value)
        if parent and normalize_text(parent) != normalize_text(header):
            candidates.append(f"{parent} {header}")
            candidates.append(f"{header} {parent}")
    return candidates


def best_vehicle_rule_from_candidates(
    candidates: Iterable[Any], rules: Iterable[VehicleRule]
) -> tuple[VehicleRule | None, float, str]:
    best_rule = None
    best_score = 0.0
    best_header = ""
    for candidate in candidates:
        rule, score = best_vehicle_rule(candidate, rules)
        if score > best_score:
            best_rule = rule
            best_score = score
            best_header = display_text(candidate)
    return best_rule, best_score, best_header


def looks_like_vehicle_header(header: Any) -> bool:
    text = display_text(header)
    if not text:
        return False
    normalized = normalize_text(text)
    if normalized in {normalize_text(alias) for aliases in ROLE_ALIASES.values() for alias in aliases}:
        return False
    return bool(VEHICLE_HINT_RE.search(text))


def detect_header_row(ws: Worksheet, rules: Iterable[VehicleRule], scan_rows: int = 20) -> int:
    best_row = 1
    best_score = -1.0
    max_row = min(ws.max_row or 1, scan_rows)

    for row_idx in range(1, max_row + 1):
        values = [cell.value for cell in ws[row_idx]]
        if not any(not is_blank(v) for v in values):
            continue

        role_scores = {role: 0.0 for role in ROLE_ALIASES}
        matched_vehicle_count = 0
        unmatched_vehicle_count = 0

        for cell in ws[row_idx]:
            candidates = header_candidates(ws, row_idx, cell.column)
            if not candidates:
                continue
            for role in ROLE_ALIASES:
                role_scores[role] = max(
                    role_scores[role],
                    *(role_match_score(candidate, role) for candidate in candidates),
                )
            _, vehicle_score, _ = best_vehicle_rule_from_candidates(candidates, rules)
            if vehicle_score:
                matched_vehicle_count += 1
            elif any(looks_like_vehicle_header(candidate) for candidate in candidates):
                unmatched_vehicle_count += 1

        required_hits = sum(1 for score in role_scores.values() if score)
        score = sum(role_scores.values()) * 2.0
        score += matched_vehicle_count * 1.5
        score += min(unmatched_vehicle_count, 3) * 0.15
        score += required_hits * 0.4

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def detect_workbook(path: str | Path, rules: list[VehicleRule] | None = None, sheet_name: str | None = None) -> DetectionResult:
    workbook_path = Path(path)
    rules = rules or load_vehicle_rules()
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row = detect_header_row(ws, rules)

    required: dict[str, DetectedColumn | None] = {role: None for role in ROLE_ALIASES}
    vehicle_columns: list[DetectedColumn] = []
    unmatched_vehicle_headers: list[DetectedColumn] = []
    duplicate_roles: dict[str, list[DetectedColumn]] = {role: [] for role in ROLE_ALIASES}

    for cell in ws[header_row]:
        base_header = display_text(cell.value)
        if not base_header:
            continue

        candidates = header_candidates(ws, header_row, cell.column)
        rule, vehicle_score, matched_header = best_vehicle_rule_from_candidates(candidates, rules)
        if rule:
            vehicle_columns.append(
                DetectedColumn(
                    role="vehicle",
                    header=matched_header or base_header,
                    index=cell.column,
                    confidence=vehicle_score,
                    rule=rule,
                    message=f"{rule.car_length} / {'、'.join(rule.vehicle_requirements)}",
                )
            )
            continue

        if any(looks_like_vehicle_header(candidate) for candidate in candidates):
            unmatched_vehicle_headers.append(
                DetectedColumn(
                    role="vehicle",
                    header=candidates[-1],
                    index=cell.column,
                    confidence=0.35,
                    rule=None,
                    message="疑似车型列，但没有匹配到规则",
                )
            )

        for role in ROLE_ALIASES:
            score = role_match_score(base_header, role)
            if score:
                duplicate_roles[role].append(
                    DetectedColumn(
                        role=role,
                        header=base_header,
                        index=cell.column,
                        confidence=score,
                    )
                )

    warnings: list[str] = []
    for role, candidates in duplicate_roles.items():
        if not candidates:
            continue
        candidates.sort(key=lambda item: (-item.confidence, item.index))
        required[role] = candidates[0]
        if len(candidates) > 1:
            extras = "、".join(f"{col.letter}:{col.header}" for col in candidates[1:])
            warnings.append(f"{ROLE_LABELS[role]} 有多个候选，已选择 {candidates[0].letter}:{candidates[0].header}；其他候选：{extras}")

    for role, label in ROLE_LABELS.items():
        if required[role] is None:
            warnings.append(f"未识别到必需字段：{label}")
    if not vehicle_columns:
        warnings.append("未识别到任何匹配车型规则的价格列")
    for col in unmatched_vehicle_headers:
        warnings.append(f"{col.letter}:{col.header} 疑似车型列，但未匹配规则，已跳过")

    return DetectionResult(
        workbook_path=workbook_path,
        sheet_name=ws.title,
        header_row=header_row,
        required_columns=required,
        vehicle_columns=vehicle_columns,
        unmatched_vehicle_headers=unmatched_vehicle_headers,
        warnings=warnings,
        max_row=ws.max_row,
    )


def create_output_path(input_path: str | Path, now: datetime | None = None) -> Path:
    path = Path(input_path)
    stamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_货拉拉报价结果_{stamp}{path.suffix}")


def value_at(ws: Worksheet, row: int, column: int) -> Any:
    cell = ws.cell(row=row, column=column)
    if cell.value is not None or not isinstance(cell, MergedCell):
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def writable_cell(ws: Worksheet, row: int, column: int):
    cell = ws.cell(row=row, column=column)
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col)
    return cell


def iter_data_rows(ws: Worksheet, detection: DetectionResult) -> list[int]:
    origin_col = detection.required_columns["origin"]
    destination_col = detection.required_columns["destination"]
    supplier_col = detection.required_columns["supplier"]
    cols = [col.index for col in (supplier_col, origin_col, destination_col) if col]
    if not cols:
        return []

    rows: list[int] = []
    for row_idx in range(detection.header_row + 1, ws.max_row + 1):
        if any(not is_blank(value_at(ws, row_idx, col_idx)) for col_idx in cols):
            rows.append(row_idx)
    return rows


def cell_has_failure(cell) -> bool:
    return bool(cell.comment and cell.comment.text.startswith(FAILURE_PREFIX))


def should_process_cell(cell, strategy: str) -> bool:
    if strategy == "all":
        return True
    if strategy == "failed":
        return cell_has_failure(cell)
    return is_blank(cell.value) and not cell_has_failure(cell)


def write_success(cell, value: float | int | str) -> None:
    cell.value = value
    cell.comment = None


def mark_failure(cell, message: str) -> None:
    clean_message = re.sub(r"\s+", " ", str(message)).strip()
    cell.comment = Comment(f"{FAILURE_PREFIX}{clean_message}", "HuolalaQuoteTool")


def numeric_or_original(value: float) -> float | int:
    if float(value).is_integer():
        return int(value)
    return value
