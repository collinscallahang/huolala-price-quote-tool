from __future__ import annotations

import hashlib
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import PriceColumn, QuoteResult, QuoteTask, WorkbookSummary
from .site_config import normalize_text, vehicle_label_for


REQUIRED_COLUMNS = {
    "供应商名称": "supplier",
    "发货地址": "origin",
    "距离": "distance",
    "到货地址": "destination",
}


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def make_file_id(path: Path) -> str:
    digest = hashlib.sha1(str(path.resolve()).encode("utf-8")).hexdigest()[:10]
    return f"{path.stem}-{digest}"


@dataclass
class WorkbookJob:
    summary: WorkbookSummary
    workbook: Any
    worksheet: Worksheet
    columns: dict[str, int]
    overwrite: bool = False

    def write_result(self, result: QuoteResult) -> None:
        if not result.success:
            if is_address_confirmation(result.error):
                self.write_address_confirmation(result)
            return
        task = self.task_for_result(result)

        if result.distance is not None and self.summary.tasks:
            distance_col = self.columns.get("distance")
            if distance_col and (self.overwrite or is_blank(self.worksheet.cell(result.row_index, distance_col).value)):
                cell = self.worksheet.cell(result.row_index, distance_col)
                cell.value = result.distance
                cell.comment = Comment(copy_comment("路程", result.distance, result.distance_source, result, task, cell.coordinate), "查价工具")

        if result.price_col and result.price is not None:
            cell = self.worksheet.cell(result.row_index, result.price_col)
            if self.overwrite or is_blank(cell.value):
                cell.value = result.price
                cell.comment = Comment(copy_comment("价格", result.price, result.price_source, result, task, cell.coordinate), "查价工具")

    def save(self) -> None:
        self.summary.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.summary.output_path)

    def task_for_result(self, result: QuoteResult) -> QuoteTask | None:
        return next((task for task in self.summary.tasks if task.task_id == result.task_id), None)

    def write_address_confirmation(self, result: QuoteResult) -> None:
        task = self.task_for_result(result)
        target_col = result.price_col or self.columns.get("distance")
        if not target_col:
            return
        cell = self.worksheet.cell(result.row_index, target_col)
        note = address_confirmation_comment(result, task, cell.coordinate)
        cell.comment = Comment(note, "查价工具")


def build_workbook_job(
    source_path: str | Path,
    output_dir: str | Path,
    length_mapping: dict[str, str] | None = None,
    overwrite: bool = False,
) -> WorkbookJob:
    source = Path(source_path)
    output_root = Path(output_dir)
    output_path = unique_output_path(output_root, f"{source.stem}_查价结果{source.suffix}")

    shutil.copy2(source, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.worksheets[0]

    header_row = find_header_row(worksheet)
    category_row = max(1, header_row - 1)
    columns = find_required_columns(worksheet, header_row)
    price_columns = find_price_columns(worksheet, category_row, header_row, columns, length_mapping)
    warnings = collect_warnings(worksheet, header_row, columns)
    file_id = make_file_id(source)

    summary = WorkbookSummary(
        file_id=file_id,
        source_path=source,
        output_path=output_path,
        sheet_name=worksheet.title,
        total_rows=max(0, worksheet.max_row - header_row),
        price_columns=price_columns,
        tasks=[],
        warnings=warnings,
    )
    job = WorkbookJob(summary=summary, workbook=workbook, worksheet=worksheet, columns=columns, overwrite=overwrite)
    summary.tasks = build_tasks(job)
    return job


def unique_output_path(output_dir: Path, filename: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    candidate = output_dir / filename
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    index = 2
    while True:
        next_candidate = output_dir / f"{stem}_{index}{suffix}"
        if not next_candidate.exists():
            return next_candidate
        index += 1


def find_header_row(worksheet: Worksheet) -> int:
    for row_index in range(1, min(worksheet.max_row, 10) + 1):
        for cell in worksheet[row_index]:
            if normalize_text(cell.value) == "供应商名称":
                return row_index
    raise ValueError("未找到表头列：供应商名称")


def find_required_columns(worksheet: Worksheet, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    reverse = {label: key for label, key in REQUIRED_COLUMNS.items()}
    for col_index in range(1, worksheet.max_column + 1):
        label = normalize_text(worksheet.cell(header_row, col_index).value)
        if label in reverse:
            columns[reverse[label]] = col_index

    missing = [label for label, key in REQUIRED_COLUMNS.items() if key not in columns]
    if missing:
        raise ValueError(f"缺少必要列：{', '.join(missing)}")
    return columns


def find_price_columns(
    worksheet: Worksheet,
    category_row: int,
    header_row: int,
    required_columns: dict[str, int],
    length_mapping: dict[str, str] | None,
) -> list[PriceColumn]:
    reserved = set(required_columns.values())
    price_columns: list[PriceColumn] = []
    for col_index in range(1, worksheet.max_column + 1):
        if col_index in reserved:
            continue
        vehicle_type = normalize_text(worksheet.cell(category_row, col_index).value)
        length_raw = normalize_text(worksheet.cell(header_row, col_index).value)
        if not vehicle_type or not length_raw:
            continue
        price_columns.append(
            PriceColumn(
                col_index=col_index,
                vehicle_type=vehicle_type,
                length_raw=length_raw,
                vehicle_label=vehicle_label_for(length_raw, length_mapping),
            )
        )
    if not price_columns:
        raise ValueError("未找到车型/价格列，请确认第 1 行和第 2 行包含车型与车长")
    return price_columns


def collect_warnings(worksheet: Worksheet, header_row: int, columns: dict[str, int]) -> list[str]:
    warnings: list[str] = []
    supplier_seen: dict[str, int] = {}
    supplier_col = columns["supplier"]
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        supplier = normalize_text(worksheet.cell(row_index, supplier_col).value)
        if not supplier:
            continue
        if supplier in supplier_seen:
            warnings.append(f"供应商重复：{supplier}，第 {supplier_seen[supplier]} 行和第 {row_index} 行会分别保留")
        else:
            supplier_seen[supplier] = row_index
    return warnings


def build_tasks(job: WorkbookJob) -> list[QuoteTask]:
    worksheet = job.worksheet
    columns = job.columns
    tasks: list[QuoteTask] = []
    supplier_col = columns["supplier"]
    origin_col = columns["origin"]
    destination_col = columns["destination"]
    distance_col = columns["distance"]
    header_row = find_header_row(worksheet)

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        supplier = normalize_text(worksheet.cell(row_index, supplier_col).value)
        origin = normalize_text(worksheet.cell(row_index, origin_col).value)
        destination = normalize_text(worksheet.cell(row_index, destination_col).value)
        if not supplier and not origin and not destination:
            continue

        needs_distance = job.overwrite or is_blank(worksheet.cell(row_index, distance_col).value)
        row_tasks: list[QuoteTask] = []
        for price_column in job.summary.price_columns:
            needs_price = job.overwrite or is_blank(worksheet.cell(row_index, price_column.col_index).value)
            if not needs_price:
                continue
            row_tasks.append(
                QuoteTask(
                    task_id=f"{job.summary.file_id}:R{row_index}C{price_column.col_index}",
                    file_id=job.summary.file_id,
                    source_file=job.summary.source_path.name,
                    row_index=row_index,
                    supplier_name=supplier,
                    origin_address=origin,
                    destination_address=destination,
                    distance_col=distance_col,
                    price_col=price_column.col_index,
                    vehicle_type=price_column.vehicle_type,
                    vehicle_length=price_column.length_raw,
                    vehicle_label=price_column.vehicle_label,
                    needs_distance=needs_distance,
                    needs_price=True,
                )
            )

        if not row_tasks and needs_distance:
            price_column = job.summary.price_columns[0]
            row_tasks.append(
                QuoteTask(
                    task_id=f"{job.summary.file_id}:R{row_index}:distance",
                    file_id=job.summary.file_id,
                    source_file=job.summary.source_path.name,
                    row_index=row_index,
                    supplier_name=supplier,
                    origin_address=origin,
                    destination_address=destination,
                    distance_col=distance_col,
                    price_col=None,
                    vehicle_type=price_column.vehicle_type,
                    vehicle_length=price_column.length_raw,
                    vehicle_label=price_column.vehicle_label,
                    needs_distance=True,
                    needs_price=False,
                )
            )

        tasks.extend(row_tasks)
    return tasks


def copy_comment(
    field_name: str,
    value: float,
    source_text: str,
    result: QuoteResult,
    task: QuoteTask | None,
    cell_coordinate: str,
) -> str:
    supplier = task.supplier_name if task else ""
    vehicle = f"{task.vehicle_label} / {task.vehicle_type}" if task else ""
    page_field = "右下角结算区「运费一口价」" if field_name == "价格" else "页面「总里程」"
    return "\n".join(
        [
            "查价复制记录",
            f"字段：{field_name}",
            f"页面字段：{page_field}",
            f"读取值：{value}",
            f"写入单元格：{cell_coordinate}",
            f"供应商：{supplier}",
            f"车型：{vehicle}",
            f"来源文本：{source_text or page_field}",
            f"任务ID：{result.task_id}",
            f"复制时间：{result.copied_at}",
        ]
    )


def is_address_confirmation(error: str) -> bool:
    return str(error or "").startswith("地址需人工确认")


def address_confirmation_comment(result: QuoteResult, task: QuoteTask | None, cell_coordinate: str) -> str:
    supplier = task.supplier_name if task else ""
    origin = task.origin_address if task else ""
    destination = task.destination_address if task else ""
    return "\n".join(
        [
            "地址需人工确认",
            f"写入单元格：{cell_coordinate}",
            f"供应商：{supplier}",
            f"发货地址：{origin}",
            f"到货地址：{destination}",
            f"原因：{result.error}",
            f"任务ID：{result.task_id}",
        ]
    )


def cell_name(row_index: int, col_index: int | None) -> str:
    if not col_index:
        return ""
    return f"{get_column_letter(col_index)}{row_index}"
