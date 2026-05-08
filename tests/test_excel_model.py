from __future__ import annotations

import unittest
import shutil
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from huolala_quote_tool.excel_model import (
    cell_has_failure,
    create_output_path,
    detect_workbook,
    load_vehicle_rules,
    mark_failure,
    should_process_cell,
    write_success,
)
from huolala_quote_tool.parsers import parse_fixed_price, parse_total_distance


class WorkspaceTempDir:
    def __enter__(self) -> str:
        self.path = Path.cwd() / f".test_tmp_{uuid.uuid4().hex}"
        self.path.mkdir()
        return str(self.path)

    def __exit__(self, exc_type, exc, tb) -> None:
        shutil.rmtree(self.path, ignore_errors=True)


def workspace_tempdir() -> WorkspaceTempDir:
    return WorkspaceTempDir()


class ExcelDetectionTests(unittest.TestCase):
    def test_detects_required_and_vehicle_columns_in_any_position(self) -> None:
        with workspace_tempdir() as tmp:
            path = Path(tmp) / "报价.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "报价单"
            ws.append(["货拉拉报价模板"])
            ws.append(["到货地址", "备注", "厢式货车&飞翼车 9.6", "供应商名称", "距离", "发货地址", "未知车型"])
            ws.append(["上海市浦东新区", "保留", None, "供应商A", None, "上海市闵行区", None])
            wb.save(path)

            detection = detect_workbook(path, rules=load_vehicle_rules())

            self.assertEqual(detection.sheet_name, "报价单")
            self.assertEqual(detection.header_row, 2)
            self.assertEqual(detection.required_columns["supplier"].index, 4)
            self.assertEqual(detection.required_columns["origin"].index, 6)
            self.assertEqual(detection.required_columns["destination"].index, 1)
            self.assertEqual(detection.required_columns["distance"].index, 5)
            self.assertEqual(len(detection.vehicle_columns), 1)
            self.assertEqual(detection.vehicle_columns[0].index, 3)
            self.assertEqual(detection.vehicle_columns[0].rule.car_length, "9米6")
            self.assertEqual(len(detection.unmatched_vehicle_headers), 1)

    def test_detects_vehicle_columns_from_parent_header_row(self) -> None:
        with workspace_tempdir() as tmp:
            path = Path(tmp) / "父表头报价.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append([None, None, None, None, "厢式货车&飞翼车", "厢式货车&飞翼车", "厢式货车&飞翼车", None])
            ws.append([None, "供应商名称", "发货地址", "距离", 9.6, "12.5&13", "16.5&17.5", "到货地址"])
            ws.append([None, "供应商A", "发货地址A", None, None, None, None, "到货地址A"])
            wb.save(path)

            detection = detect_workbook(path, rules=load_vehicle_rules())

            self.assertEqual(detection.header_row, 2)
            self.assertEqual(len(detection.vehicle_columns), 3)
            self.assertEqual([col.rule.car_length for col in detection.vehicle_columns], ["9米6", "13米", "17米5"])
            self.assertFalse(detection.unmatched_vehicle_headers)

    def test_loads_external_csv_rules_and_skips_disabled_rows(self) -> None:
        with workspace_tempdir() as tmp:
            path = Path(tmp) / "vehicle_rules.csv"
            path.write_text(
                "规则名称,Excel表头关键词,网页车长,网页车型要求,启用\n"
                "可用规则,测试车型9.6,9米6,厢式货车;飞翼车,是\n"
                "停用规则,停用车型9.6,9米6,平板车,否\n",
                encoding="utf-8",
            )

            rules = load_vehicle_rules(path)

            self.assertEqual(len(rules), 1)
            self.assertEqual(rules[0].name, "可用规则")
            self.assertEqual(rules[0].excel_headers, ("测试车型9.6",))
            self.assertEqual(rules[0].vehicle_requirements, ("厢式货车", "飞翼车"))

    def test_output_path_does_not_overwrite_original(self) -> None:
        result = create_output_path(Path("demo.xlsx"), now=datetime(2026, 5, 9, 10, 11, 12))
        self.assertEqual(result.name, "demo_货拉拉报价结果_20260509_101112.xlsx")

    def test_write_back_keeps_workbook_formatting(self) -> None:
        with workspace_tempdir() as tmp:
            path = Path(tmp) / "source.xlsx"
            output = Path(tmp) / "result.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.column_dimensions["C"].width = 28
            ws["C2"] = None
            ws["C2"].font = Font(bold=True)
            wb.save(path)

            wb2 = load_workbook(path)
            ws2 = wb2.active
            write_success(ws2["C2"], 1072.64)
            wb2.save(output)

            saved = load_workbook(output)
            self.assertEqual(saved.active["C2"].value, 1072.64)
            self.assertEqual(saved.active.column_dimensions["C"].width, 28)
            self.assertTrue(saved.active["C2"].font.bold)

    def test_failure_comments_drive_rerun_strategy(self) -> None:
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        self.assertTrue(should_process_cell(cell, "empty"))
        mark_failure(cell, "页面未生成运费一口价")
        self.assertTrue(cell_has_failure(cell))
        self.assertFalse(should_process_cell(cell, "empty"))
        self.assertTrue(should_process_cell(cell, "failed"))
        self.assertTrue(should_process_cell(cell, "all"))


class ParserTests(unittest.TestCase):
    def test_parse_total_distance_and_fixed_price(self) -> None:
        text = "总里程145公里\n运费一口价 1072.64元\n总计 992.64元"
        self.assertEqual(parse_total_distance(text), 145)
        self.assertEqual(parse_fixed_price(text), 1072.64)


if __name__ == "__main__":
    unittest.main()
