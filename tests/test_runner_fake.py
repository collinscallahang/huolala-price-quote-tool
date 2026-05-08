from __future__ import annotations

import unittest
import uuid
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from price_quote_tool.models import QuoteResult
from price_quote_tool.runner import BatchRun


TEST_TMP = Path(__file__).resolve().parents[1] / ".test_tmp"


def make_case_dir() -> Path:
    path = TEST_TMP / f"case_{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    return path


class FakeQuoteClient:
    def open(self) -> None:
        pass

    def close(self) -> None:
        pass

    def quote(self, task):
        price_by_label = {"9米6": 1072.64, "13米": 1472.67, "17米5": 1731.53}
        return QuoteResult(
            task_id=task.task_id,
            file_id=task.file_id,
            row_index=task.row_index,
            price_col=task.price_col,
            success=True,
            price=price_by_label.get(task.vehicle_label, 999.0),
            distance=145,
        )


class RowFakeQuoteClient:
    def __init__(self) -> None:
        self.row_calls = 0

    def open(self) -> None:
        pass

    def close(self) -> None:
        pass

    def quote_row(self, tasks):
        self.row_calls += 1
        prices = {"9米6": 100, "13米": 200, "17米5": 300}
        results = []
        for index, task in enumerate(tasks):
            results.append(
                QuoteResult(
                    task_id=task.task_id,
                    file_id=task.file_id,
                    row_index=task.row_index,
                    price_col=task.price_col,
                    success=True,
                    price=prices[task.vehicle_label],
                    distance=88 if index == 0 else None,
                    price_source=f"运费一口价 {task.vehicle_label}",
                    distance_source="总里程88公里" if index == 0 else "",
                    copied_at="2026-05-08T10:00:00",
                )
            )
        return results


class RunnerTests(unittest.TestCase):
    def test_batch_run_writes_one_output_per_input(self) -> None:
        root = make_case_dir()
        config_dir = root / "configs"
        config_dir.mkdir()
        config = config_dir / "site.huolala.json"
        output_root = root / "configured_output"
        config.write_text(
            json.dumps(
                {
                    "vehicle_length_mapping": {"9.6": "9米6"},
                    "output_root": str(output_root),
                    "workflow": [],
                    "extract": {},
                    "browser": {},
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        sample = root / "样本.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append([None, None, None, None, "厢式货车&飞翼门", None])
        ws.append([None, "供应商名称", "发货地址", "距离", 9.6, "到货地址"])
        ws.append([None, "供应商A", "合肥A路1号", None, None, "滁州B路2号"])
        wb.save(sample)

        run = BatchRun(
            run_id="test_run",
            excel_paths=[sample],
            site_url="http://example.test",
            retry_count=1,
            root_dir=root,
            config_path=config,
            quote_client_factory=FakeQuoteClient,
        )
        run._run()

        self.assertEqual(run.progress.status, "completed")
        self.assertEqual(len(run.result_files()), 1)
        self.assertEqual(run.output_dir, output_root / "test_run")
        result_wb = load_workbook(run.result_files()[0], data_only=True)
        result_ws = result_wb.active
        self.assertEqual(result_ws.cell(3, 4).value, 145)
        self.assertEqual(result_ws.cell(3, 5).value, 1072.64)

    def test_batch_run_groups_same_row_and_writes_copy_records(self) -> None:
        root = make_case_dir()
        config_dir = root / "configs"
        config_dir.mkdir()
        config = config_dir / "site.huolala.json"
        output_root = root / "configured_output"
        config.write_text(
            json.dumps(
                {
                    "vehicle_length_mapping": {
                        "9.6": "9米6",
                        "12.5&13": "13米",
                        "16.5&17.5": "17米5",
                    },
                    "output_root": str(output_root),
                    "workflow": [],
                    "extract": {},
                    "browser": {},
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        sample = root / "样本.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append([None, None, None, None, "厢式货车&飞翼门", "厢式货车&飞翼门", "厢式货车&飞翼门", None])
        ws.append([None, "供应商名称", "发货地址", "距离", 9.6, "12.5&13", "16.5&17.5", "到货地址"])
        ws.append([None, "供应商A", "合肥A路1号", None, None, None, None, "滁州B路2号"])
        wb.save(sample)

        client = RowFakeQuoteClient()
        run = BatchRun(
            run_id="row_test",
            excel_paths=[sample],
            site_url="http://example.test",
            retry_count=1,
            root_dir=root,
            config_path=config,
            quote_client_factory=lambda: client,
        )
        run._run()

        self.assertEqual(client.row_calls, 1)
        result_wb = load_workbook(run.result_files()[0], data_only=True)
        result_ws = result_wb.active
        self.assertEqual(result_ws.cell(3, 4).value, 88)
        self.assertEqual([result_ws.cell(3, col).value for col in (5, 6, 7)], [100, 200, 300])
        self.assertIsNotNone(result_ws.cell(3, 5).comment)
        self.assertTrue((run.output_dir / "复制粘贴记录.csv").exists())


if __name__ == "__main__":
    unittest.main()
