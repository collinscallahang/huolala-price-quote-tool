from __future__ import annotations

import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook

from price_quote_tool.automation import settlement_price
from price_quote_tool.excel_io import build_workbook_job
from price_quote_tool.models import QuoteResult
from price_quote_tool.site_config import vehicle_requirement_labels_for


TEST_TMP = Path(__file__).resolve().parents[1] / ".test_tmp"


def make_case_dir() -> Path:
    path = TEST_TMP / f"case_{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    return path


class ExcelIoTests(unittest.TestCase):
    def make_sample(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append([None, None, None, None, "厢式货车&飞翼门", "厢式货车&飞翼门", "厢式货车&飞翼门", None])
        ws.append([None, "供应商名称", "发货地址", "距离", 9.6, "12.5&13", "16.5&17.5", "到货地址"])
        ws.append([None, "供应商A", "合肥A路1号", None, None, None, None, "滁州B路2号"])
        ws.append([None, "供应商B", "合肥C路3号", 88, 100, None, 300, "滁州D路4号"])
        wb.save(path)

    def test_detect_two_row_headers_and_only_blank_tasks(self) -> None:
        tmp = make_case_dir()
        source = tmp / "样本.xlsx"
        output = tmp / "out"
        self.make_sample(source)

        job = build_workbook_job(source, output)

        self.assertEqual([c.vehicle_label for c in job.summary.price_columns], ["9米6", "13米", "17米5"])
        self.assertEqual(len(job.summary.tasks), 4)
        self.assertTrue(all(task.source_file == "样本.xlsx" for task in job.summary.tasks))

    def test_write_result_preserves_existing_values(self) -> None:
        tmp = make_case_dir()
        source = tmp / "样本.xlsx"
        output = tmp / "out"
        self.make_sample(source)

        job = build_workbook_job(source, output)
        first = job.summary.tasks[0]
        job.write_result(
            QuoteResult(
                task_id=first.task_id,
                file_id=first.file_id,
                row_index=first.row_index,
                price_col=first.price_col,
                success=True,
                price=1072.64,
                distance=145,
            )
        )
        second_row_task = [task for task in job.summary.tasks if task.row_index == 4][0]
        job.write_result(
            QuoteResult(
                task_id=second_row_task.task_id,
                file_id=second_row_task.file_id,
                row_index=second_row_task.row_index,
                price_col=second_row_task.price_col,
                success=True,
                price=1472.67,
                distance=999,
            )
        )
        job.save()

        wb = load_workbook(job.summary.output_path, data_only=True)
        ws = wb.active
        self.assertEqual(ws.cell(3, 4).value, 145)
        self.assertEqual(ws.cell(3, 5).value, 1072.64)
        self.assertEqual(ws.cell(4, 4).value, 88)
        self.assertEqual(ws.cell(4, 5).value, 100)
        self.assertEqual(ws.cell(4, 6).value, 1472.67)

    def test_vehicle_requirements_follow_excel_tokens(self) -> None:
        mapping = {
            "厢式货车": ["厢式货车"],
            "飞翼门": ["飞翼车"],
        }

        self.assertEqual(
            vehicle_requirement_labels_for("厢式货车&飞翼门", mapping),
            ["厢式货车", "飞翼车"],
        )
        self.assertEqual(vehicle_requirement_labels_for("平板货车", mapping), ["平板货车"])

    def test_settlement_price_prefers_bottom_freight_price(self) -> None:
        body_text = """
特快
一口价 739.62元
快车
一口价 672.38元
顺路车
一口价 537.9元
货运券抵扣
-80元
一口价
运费一口价 672.38元
总计
592.38元
总里程47公里
下一步
"""
        patterns = [r"运费\s*[—\-－]?\s*一口价\s*([0-9,]+(?:\.[0-9]+)?)\s*元"]

        self.assertEqual(settlement_price(body_text, patterns), 672.38)


if __name__ == "__main__":
    unittest.main()
